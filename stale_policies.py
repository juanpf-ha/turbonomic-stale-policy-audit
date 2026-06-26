#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Turbonomic Stale Policy Audit - VMware On-Premises Focus

Read-only audit script for Turbonomic policies. It is intentionally conservative:
missing audit/action data is not treated as proof that a policy is stale.

Main data sources:
  - /api/v3/targets
  - /api/v3/search
  - /api/v3/audit                 (optional; can be unavailable)
  - /api/v3/settingspolicies      (automation/settings policies)
  - /api/v3/settingspolicies?only_defaults=true
  - /api/v3/policies              (placement/policy endpoint, not automation policies)
  - /api/v3/actions               (best-effort; version-dependent)
  - /api/v3/groups/{uuid}         (best-effort scope validation)

The script does not create, modify, delete, or execute anything in Turbonomic.
"""

from __future__ import annotations

import argparse
import copy
import hashlib
import json
import os
import re
import sys
from collections import defaultdict
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Iterable

import requests
import urllib3
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


# -----------------------------------------------------------------------------
# CLI
# -----------------------------------------------------------------------------

parser = argparse.ArgumentParser(
    description="Turbonomic Stale Policy Audit - VMware On-Premises Focus"
)
parser.add_argument("--host", required=True, help="Turbonomic host URL")
parser.add_argument("--user", required=True, help="API username")
parser.add_argument("--password", required=True, help="API password")
parser.add_argument("--days", default=90, type=int, help="Inactivity threshold in days")
parser.add_argument(
    "--output",
    default="vmware_stale_policies.xlsx",
    help="Output Excel file. Relative paths are written under the current directory.",
)
parser.add_argument("--verbose", action="store_true", help="Enable verbose output")
parser.add_argument(
    "--snapshot-dir",
    default=None,
    help="Directory for JSON snapshots. Default: <output-dir>/snapshots",
)
parser.add_argument(
    "--no-snapshot",
    action="store_true",
    help="Do not write or compare JSON snapshots",
)
parser.add_argument(
    "--skip-group-check",
    action="store_true",
    help="Do not call /api/v3/groups/{uuid} for scope validation",
)
parser.add_argument(
    "--skip-action-check",
    action="store_true",
    help="Do not call /api/v3/actions for best-effort recent action checks",
)
parser.add_argument(
    "--trust-env-proxy",
    action="store_true",
    help="Allow requests to use proxy variables from the environment. Default: disabled.",
)

args = parser.parse_args()

TURBO_HOST = args.host.rstrip("/")
USERNAME = args.user
PASSWORD = args.password
INACTIVITY_DAYS = args.days
OUTPUT_FILE = args.output
VERBOSE = args.verbose
SKIP_GROUP_CHECK = args.skip_group_check
SKIP_ACTION_CHECK = args.skip_action_check


# -----------------------------------------------------------------------------
# Constants
# -----------------------------------------------------------------------------

# VMware-relevant Turbonomic entity types. In the API, ESXi hosts are typically
# represented as PhysicalMachine and vCenter datastores as Storage.
VMWARE_ENTITY_TYPES = [
    "VirtualMachine",
    "PhysicalMachine",
    "Storage",
    "Datacenter",
    "Cluster",
    "VirtualDataCenter",
    "ResourcePool",
]

TARGET_OK_STATUSES = {"validated", "discovered"}

TEST_NAME_PATTERNS = [
    "test",
    "demo",
    "temp",
    "tmp",
    "sandbox",
    "prueba",
    "ejemplo",
    "poc",
    "pilot",
    "trial",
    "dev-",
    "qa-",
    "staging-",
]

ORPHAN_PATTERNS = [
    "::deleted",
    "::removed",
    "::migrated",
    "::decommissioned",
    "old-vcenter",
    "legacy-",
    "retired-",
    "deprecated",
]

DEFAULT_POLICY_PATTERNS = [
    "default",
    "defaults",
    "system default",
    "turbonomic default",
]

ACTION_MODE_VALUES = {
    "DISABLED",
    "RECOMMEND",
    "MANUAL",
    "AUTOMATIC",
    "EXTERNAL_APPROVAL",
}

ACTION_MODE_RISK = {
    "AUTOMATIC": "HIGH",
    "EXTERNAL_APPROVAL": "MEDIUM",
    "MANUAL": "MEDIUM",
    "RECOMMEND": "LOW",
    "DISABLED": "INFO",
}

CLASSIFICATION_ORDER = [
    "CANDIDATE_DELETE",
    "REVIEW",
    "KEEP",
    "INFO",
    "UNKNOWN",
]

CLASSIFICATION_COLORS = {
    "CANDIDATE_DELETE": "FFCCCC",
    "REVIEW": "FFF2CC",
    "KEEP": "CCFFCC",
    "INFO": "D9EAF7",
    "UNKNOWN": "E7E6E6",
}


# -----------------------------------------------------------------------------
# Generic helpers
# -----------------------------------------------------------------------------


def now_utc() -> datetime:
    return datetime.now(timezone.utc)


def debug(message: str) -> None:
    if VERBOSE:
        print(message)


def normalize_string(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def lower(value: Any) -> str:
    return normalize_string(value).lower()


def sanitize_sheet_name(name: str) -> str:
    cleaned = re.sub(r"[\\/*?:\[\]]", "_", name)
    return cleaned[:31]


def resolve_output_path(output_file: str) -> Path:
    path = Path(output_file).expanduser()
    if not path.is_absolute():
        path = Path.cwd() / path
    path.parent.mkdir(parents=True, exist_ok=True)
    return path


def unwrap_list_response(payload: Any) -> list[Any]:
    """Return a list from API responses that may be a list or wrapped dict."""
    if isinstance(payload, list):
        return payload
    if isinstance(payload, dict):
        for key in (
            "results",
            "content",
            "items",
            "data",
            "targetList",
            "targets",
            "policies",
            "settingsPolicies",
        ):
            value = payload.get(key)
            if isinstance(value, list):
                return value
    return []


def json_hash(obj: Any) -> str:
    raw = json.dumps(obj, sort_keys=True, ensure_ascii=False, default=str)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def api_get(session: requests.Session, endpoint: str, **kwargs: Any) -> requests.Response:
    return session.get(f"{TURBO_HOST}{endpoint}", **kwargs)


def api_post(session: requests.Session, endpoint: str, **kwargs: Any) -> requests.Response:
    return session.post(f"{TURBO_HOST}{endpoint}", **kwargs)


def print_api_warning(label: str, response: requests.Response) -> None:
    print(f"  Unable to retrieve {label}: HTTP {response.status_code}")
    text = response.text or ""
    if text:
        print(f"  Response: {text[:500]}")


# -----------------------------------------------------------------------------
# Target helpers
# -----------------------------------------------------------------------------


def get_target_field(target: dict[str, Any], field_name: str) -> Any:
    for field in target.get("inputFields", []) or []:
        if isinstance(field, dict) and field.get("name") == field_name:
            return field.get("value")
    return None


def is_vmware_target(target: dict[str, Any]) -> bool:
    """
    Detect VMware/vCenter targets.

    Turbonomic commonly reports vCenter as category=Hypervisor and type=vCenter,
    not as type=VMware.
    """
    category = lower(target.get("category"))
    type_name = lower(target.get("type"))
    display_name = lower(target.get("displayName") or target.get("name"))
    address = lower(
        get_target_field(target, "address")
        or get_target_field(target, "nameOrAddress")
        or target.get("address")
    )

    if category == "hypervisor" and type_name in {"vcenter", "vmware", "vsphere", "vmware vcenter"}:
        return True

    haystack = " ".join([type_name, display_name, address])
    return any(token in haystack for token in ("vcenter", "vmware", "vsphere"))


def is_target_ok(status: Any) -> bool:
    return lower(status) in TARGET_OK_STATUSES


# -----------------------------------------------------------------------------
# Policy helpers
# -----------------------------------------------------------------------------


def policy_name(policy: dict[str, Any]) -> str:
    return normalize_string(
        policy.get("displayName")
        or policy.get("name")
        or policy.get("policyName")
        or policy.get("uuid")
        or "N/A"
    )


def policy_id(policy: dict[str, Any]) -> str:
    return normalize_string(policy.get("uuid") or policy.get("id") or policy.get("policyUuid"))


def policy_entity_type(policy: dict[str, Any]) -> str:
    return normalize_string(policy.get("entityType") or policy.get("entity_type") or "UNKNOWN")


def policy_enabled(policy: dict[str, Any]) -> bool:
    """Turbonomic settings policies usually use disabled=false, not enabled=true."""
    if "enabled" in policy:
        return bool(policy.get("enabled"))
    if "disabled" in policy:
        return not bool(policy.get("disabled"))
    return True


def policy_scopes(policy: dict[str, Any]) -> list[Any]:
    scopes = policy.get("scopes", policy.get("scope", []))
    if scopes is None:
        return []
    if isinstance(scopes, list):
        return scopes
    if isinstance(scopes, dict):
        return [scopes]
    return []


def extract_ref_uuid(ref: Any) -> str:
    if isinstance(ref, str):
        return ref
    if isinstance(ref, dict):
        for key in ("uuid", "id", "groupUuid", "targetId"):
            if ref.get(key):
                return str(ref[key])
    return ""


def extract_ref_name(ref: Any) -> str:
    if isinstance(ref, dict):
        return normalize_string(ref.get("displayName") or ref.get("name") or ref.get("uuid") or ref.get("id"))
    return normalize_string(ref)


def scope_ids(policy: dict[str, Any]) -> list[str]:
    ids = []
    for ref in policy_scopes(policy):
        ref_id = extract_ref_uuid(ref)
        if ref_id:
            ids.append(ref_id)
    return sorted(set(ids))


def is_default_policy(policy: dict[str, Any], default_ids: set[str] | None = None) -> bool:
    pid = policy_id(policy)
    name = lower(policy_name(policy))
    if default_ids and pid in default_ids:
        return True
    if bool(policy.get("isDefault")) or bool(policy.get("default")):
        return True
    if name.endswith(" defaults") or name.endswith("defaults"):
        return True
    return any(pattern in name for pattern in DEFAULT_POLICY_PATTERNS)


def has_name_pattern(name: str, patterns: Iterable[str]) -> bool:
    name_l = lower(name)
    return any(pattern in name_l for pattern in patterns)


def parse_last_modified(value: Any) -> datetime | None:
    if not value:
        return None
    if isinstance(value, (int, float)):
        # Turbonomic often uses epoch milliseconds.
        if value > 10_000_000_000:
            return datetime.fromtimestamp(value / 1000, tz=timezone.utc)
        return datetime.fromtimestamp(value, tz=timezone.utc)
    text = str(value).strip()
    if not text:
        return None
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00"))
    except ValueError:
        return None


def policy_age_days(policy: dict[str, Any]) -> int | None:
    dt = parse_last_modified(policy.get("lastModified") or policy.get("modifiedTime") or policy.get("updateTime"))
    if not dt:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return (now_utc() - dt).days


# -----------------------------------------------------------------------------
# Settings flattening and default comparison
# -----------------------------------------------------------------------------


def flatten_policy_settings(policy: dict[str, Any]) -> dict[str, dict[str, Any]]:
    """
    Best-effort extraction of policy settings.

    Turbonomic SettingsPolicyApiDTOs can contain settingsManagers with nested
    settings. This function intentionally walks the structure generically so it
    works across versions.
    """
    flattened: dict[str, dict[str, Any]] = {}
    managers = policy.get("settingsManagers") or policy.get("settingManagers") or []
    if not isinstance(managers, list):
        managers = []

    def walk(obj: Any, path: str, manager_name: str, manager_uuid: str) -> None:
        if isinstance(obj, dict):
            setting_uuid = normalize_string(
                obj.get("uuid")
                or obj.get("id")
                or obj.get("name")
                or obj.get("settingName")
                or obj.get("displayName")
            )
            has_value = any(k in obj for k in ("value", "valueType", "defaultValue"))
            if setting_uuid and has_value:
                value = obj.get("value")
                display = normalize_string(obj.get("displayName") or obj.get("name") or setting_uuid)
                key = f"{manager_uuid or manager_name}|{setting_uuid}"
                flattened[key] = {
                    "key": key,
                    "manager": manager_name,
                    "manager_uuid": manager_uuid,
                    "setting_uuid": setting_uuid,
                    "display_name": display,
                    "value": value,
                    "value_type": obj.get("valueType"),
                    "default_value": obj.get("defaultValue"),
                    "path": path,
                }

            for k, v in obj.items():
                if k in {"links"}:
                    continue
                walk(v, f"{path}.{k}" if path else k, manager_name, manager_uuid)
        elif isinstance(obj, list):
            for idx, item in enumerate(obj):
                walk(item, f"{path}[{idx}]", manager_name, manager_uuid)

    for index, manager in enumerate(managers):
        if not isinstance(manager, dict):
            continue
        manager_name = normalize_string(manager.get("displayName") or manager.get("name") or f"manager_{index}")
        manager_uuid = normalize_string(manager.get("uuid") or manager.get("id") or manager_name)
        walk(manager, f"settingsManagers[{index}]", manager_name, manager_uuid)

    return flattened


def build_default_policy_maps(default_policies: list[dict[str, Any]]) -> tuple[dict[str, dict[str, Any]], dict[str, dict[str, dict[str, Any]]], set[str]]:
    default_by_entity: dict[str, dict[str, Any]] = {}
    default_settings_by_entity: dict[str, dict[str, dict[str, Any]]] = {}
    default_ids: set[str] = set()

    for policy in default_policies:
        pid = policy_id(policy)
        etype = policy_entity_type(policy)
        if pid:
            default_ids.add(pid)
        if etype and etype not in default_by_entity:
            default_by_entity[etype] = policy
            default_settings_by_entity[etype] = flatten_policy_settings(policy)

    return default_by_entity, default_settings_by_entity, default_ids


def compare_to_default(
    policy: dict[str, Any],
    default_settings_by_entity: dict[str, dict[str, dict[str, Any]]],
) -> dict[str, Any]:
    etype = policy_entity_type(policy)
    settings = flatten_policy_settings(policy)
    default_settings = default_settings_by_entity.get(etype)

    if not settings:
        return {
            "status": "NO_SETTINGS",
            "changed_count": 0,
            "total_settings": 0,
            "changed_keys": [],
        }

    if not default_settings:
        return {
            "status": "NO_DEFAULT_AVAILABLE",
            "changed_count": None,
            "total_settings": len(settings),
            "changed_keys": [],
        }

    changed_keys = []
    for key, setting in settings.items():
        default_setting = default_settings.get(key)
        if not default_setting:
            changed_keys.append(key)
            continue
        if setting.get("value") != default_setting.get("value"):
            changed_keys.append(key)

    return {
        "status": "COMPARED",
        "changed_count": len(changed_keys),
        "total_settings": len(settings),
        "changed_keys": changed_keys[:25],
    }


def extract_action_modes(policy: dict[str, Any]) -> list[dict[str, Any]]:
    modes = []
    settings = flatten_policy_settings(policy)
    for item in settings.values():
        value = normalize_string(item.get("value")).upper()
        if value in ACTION_MODE_VALUES:
            modes.append(
                {
                    "policy_id": policy_id(policy),
                    "policy_name": policy_name(policy),
                    "entity_type": policy_entity_type(policy),
                    "manager": item.get("manager"),
                    "setting_uuid": item.get("setting_uuid"),
                    "setting_name": item.get("display_name"),
                    "value": value,
                    "risk": ACTION_MODE_RISK.get(value, "UNKNOWN"),
                }
            )
    return modes


# -----------------------------------------------------------------------------
# Scope and group validation
# -----------------------------------------------------------------------------


def extract_member_count(group: dict[str, Any]) -> int | None:
    for key in (
        "memberCount",
        "membersCount",
        "numMembers",
        "entitiesCount",
        "entityCount",
        "count",
    ):
        value = group.get(key)
        if isinstance(value, int):
            return value
    return None


def fetch_group(session: requests.Session, group_uuid: str, cache: dict[str, dict[str, Any]]) -> dict[str, Any]:
    if group_uuid in cache:
        return cache[group_uuid]

    result = {
        "uuid": group_uuid,
        "resolved": False,
        "status_code": None,
        "display_name": "",
        "group_type": "",
        "class_name": "",
        "member_count": None,
        "error": "",
    }

    if SKIP_GROUP_CHECK:
        result["error"] = "group check skipped"
        cache[group_uuid] = result
        return result

    try:
        response = api_get(session, f"/api/v3/groups/{group_uuid}")
        result["status_code"] = response.status_code
        if response.status_code == 200:
            payload = response.json()
            if isinstance(payload, dict):
                result.update(
                    {
                        "resolved": True,
                        "display_name": normalize_string(payload.get("displayName") or payload.get("name")),
                        "group_type": normalize_string(payload.get("groupType") or payload.get("entityType")),
                        "class_name": normalize_string(payload.get("className") or payload.get("groupClassName")),
                        "member_count": extract_member_count(payload),
                    }
                )
        else:
            result["error"] = response.text[:300]
    except Exception as exc:  # noqa: BLE001 - best-effort diagnostic
        result["error"] = str(exc)

    cache[group_uuid] = result
    return result


def analyze_scopes(
    session: requests.Session,
    policy: dict[str, Any],
    default_policy: bool,
    group_cache: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    refs = policy_scopes(policy)
    ids = scope_ids(policy)

    if not refs:
        if default_policy:
            return {
                "scope_health": "DEFAULT_OR_GLOBAL",
                "scope_count": 0,
                "resolved_count": 0,
                "empty_group_count": 0,
                "unresolved_count": 0,
                "scope_details": [],
            }
        return {
            "scope_health": "EMPTY",
            "scope_count": 0,
            "resolved_count": 0,
            "empty_group_count": 0,
            "unresolved_count": 0,
            "scope_details": [],
        }

    details = []
    resolved = 0
    unresolved = 0
    empty_groups = 0

    for ref in refs:
        gid = extract_ref_uuid(ref)
        ref_name = extract_ref_name(ref)
        if not gid:
            unresolved += 1
            details.append(
                {
                    "uuid": "",
                    "name": ref_name,
                    "resolved": False,
                    "member_count": None,
                    "health": "UNRESOLVED",
                    "error": "scope reference has no uuid/id",
                }
            )
            continue

        group = fetch_group(session, gid, group_cache)
        if group.get("resolved"):
            resolved += 1
            member_count = group.get("member_count")
            if member_count == 0:
                empty_groups += 1
            details.append(
                {
                    "uuid": gid,
                    "name": group.get("display_name") or ref_name,
                    "resolved": True,
                    "member_count": member_count,
                    "group_type": group.get("group_type"),
                    "class_name": group.get("class_name"),
                    "health": "EMPTY_GROUP" if member_count == 0 else "OK",
                    "error": "",
                }
            )
        else:
            unresolved += 1
            details.append(
                {
                    "uuid": gid,
                    "name": ref_name,
                    "resolved": False,
                    "member_count": None,
                    "health": "UNRESOLVED",
                    "error": group.get("error") or f"HTTP {group.get('status_code')}",
                }
            )

    if unresolved == len(refs):
        health = "UNRESOLVED"
    elif empty_groups > 0:
        health = "EMPTY_GROUP"
    elif unresolved > 0:
        health = "PARTIAL"
    else:
        health = "OK"

    return {
        "scope_health": health,
        "scope_count": len(refs),
        "resolved_count": resolved,
        "empty_group_count": empty_groups,
        "unresolved_count": unresolved,
        "scope_details": details,
        "scope_ids": ids,
    }


# -----------------------------------------------------------------------------
# API collection
# -----------------------------------------------------------------------------


def login() -> requests.Session:
    session = requests.Session()
    session.verify = False
    session.trust_env = bool(args.trust_env_proxy)

    print(f"\nConnecting to {TURBO_HOST} ...")
    response = api_post(
        session,
        "/api/v3/login",
        data={"username": USERNAME, "password": PASSWORD},
    )
    if response.status_code != 200:
        print(f"Authentication failed: HTTP {response.status_code}")
        if response.text:
            print(response.text[:500])
        raise SystemExit(1)

    print("Authentication successful")
    return session


def analyze_vmware_targets(session: requests.Session) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    print("\n-- VMware vCenter Targets ---------------------------")
    try:
        response = api_get(session, "/api/v3/targets")
        if response.status_code != 200:
            print_api_warning("targets", response)
            return [], []

        targets = unwrap_list_response(response.json())
        vmware_targets = [t for t in targets if isinstance(t, dict) and is_vmware_target(t)]
        print(f"  Total VMware targets: {len(vmware_targets)}")

        disconnected = []
        for target in vmware_targets:
            status = target.get("status", "UNKNOWN")
            name = normalize_string(
                target.get("displayName")
                or target.get("name")
                or get_target_field(target, "address")
                or target.get("uuid")
                or "N/A"
            )
            if is_target_ok(status):
                debug(f"  OK: {name}: {status}")
            else:
                disconnected.append({"name": name, "status": status, "uuid": target.get("uuid")})
                print(f"  WARNING: {name}: {status}")

        if vmware_targets and not disconnected:
            print("  All VMware/vCenter targets are in an accepted state")

        return vmware_targets, disconnected
    except Exception as exc:  # noqa: BLE001 - diagnostic script
        print(f"  Error analyzing VMware targets: {exc}")
        return [], []


def get_vmware_entity_counts(session: requests.Session) -> dict[str, int | str]:
    print("\n-- VMware Entity Inventory --------------------------")
    entity_counts: dict[str, int | str] = {}

    for entity_type in VMWARE_ENTITY_TYPES:
        try:
            response = api_get(
                session,
                "/api/v3/search",
                params={"types": entity_type, "limit": 1},
            )
            if response.status_code == 200:
                payload = response.json()
                if isinstance(payload, dict) and isinstance(payload.get("count"), int):
                    count: int | str = payload["count"]
                else:
                    # Some versions return only the page; with limit=1 this is an
                    # availability indicator rather than a true total.
                    count = len(unwrap_list_response(payload))
                entity_counts[entity_type] = count
                debug(f"  {entity_type}: {count}")
            else:
                entity_counts[entity_type] = "N/A"
                debug(f"  {entity_type}: HTTP {response.status_code}")
        except Exception as exc:  # noqa: BLE001
            entity_counts[entity_type] = "N/A"
            debug(f"  {entity_type}: {exc}")

    numeric_total = sum(v for v in entity_counts.values() if isinstance(v, int))
    print(f"  Total VMware entities: {numeric_total}")
    return entity_counts


def collect_audit_log(session: requests.Session, cutoff_ms: int) -> tuple[bool, set[str], int, str]:
    print(f"\n-- Audit Log (last {INACTIVITY_DAYS} days) --------------------------")
    policy_ids_in_audit: set[str] = set()
    audit_count = 0
    note = ""

    try:
        offset = 0
        limit = 500
        while True:
            response = api_get(
                session,
                "/api/v3/audit",
                params={"starttime": cutoff_ms, "limit": limit, "offset": offset},
            )
            if response.status_code != 200:
                note = f"unavailable: HTTP {response.status_code}"
                print(f"  Audit log unavailable: {response.status_code}")
                return False, policy_ids_in_audit, audit_count, note

            entries = unwrap_list_response(response.json())
            if not entries:
                break

            audit_count += len(entries)
            for entry in entries:
                if not isinstance(entry, dict):
                    continue
                for key in ("targetId", "policyId", "uuid", "entityUuid"):
                    value = entry.get(key)
                    if value:
                        policy_ids_in_audit.add(str(value))

            if len(entries) < limit:
                break
            offset += limit

        print(f"  Audit entries retrieved: {audit_count}")
        print(f"  Policies/entities with activity: {len(policy_ids_in_audit)}")
        return True, policy_ids_in_audit, audit_count, "available"
    except Exception as exc:  # noqa: BLE001
        note = f"error: {exc}"
        print(f"  Audit log error: {exc}")
        return False, policy_ids_in_audit, audit_count, note


def get_recent_policy_action_status(
    session: requests.Session,
    pid: str,
    cutoff_ms: int,
    cache: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    """
    Best-effort check. Some Turbonomic versions may not support policyId on
    /actions. A failed call is recorded as N/A and is not used as stale evidence.
    """
    if SKIP_ACTION_CHECK:
        return {"available": False, "has_recent_actions": None, "status_code": None, "note": "skipped"}
    if not pid:
        return {"available": False, "has_recent_actions": None, "status_code": None, "note": "no policy id"}
    if pid in cache:
        return cache[pid]

    result = {"available": False, "has_recent_actions": None, "status_code": None, "note": ""}
    try:
        response = api_get(
            session,
            "/api/v3/actions",
            params={"policyId": pid, "starttime": cutoff_ms, "limit": 1},
        )
        result["status_code"] = response.status_code
        if response.status_code == 200:
            payload = response.json()
            entries = unwrap_list_response(payload)
            result["available"] = True
            result["has_recent_actions"] = bool(entries)
            result["note"] = "available"
        else:
            result["note"] = f"HTTP {response.status_code}"
    except Exception as exc:  # noqa: BLE001
        result["note"] = f"error: {exc}"

    cache[pid] = result
    return result


def collect_settings_policies(session: requests.Session) -> tuple[list[dict[str, Any]], str]:
    print("\n-- Automation / Settings Policies -------------------")
    response = api_get(session, "/api/v3/settingspolicies")
    if response.status_code == 200:
        policies = [p for p in unwrap_list_response(response.json()) if isinstance(p, dict)]
        print(f"  Total settings policies: {len(policies)}")
        vmware_count = sum(1 for p in policies if policy_entity_type(p) in VMWARE_ENTITY_TYPES)
        print(f"  VMware-specific: {vmware_count}")
        return policies, "available"

    print_api_warning("settings policies", response)
    return [], f"HTTP {response.status_code}"


def collect_default_settings_policies(session: requests.Session) -> tuple[list[dict[str, Any]], str]:
    print("\n-- Default Automation Policies ----------------------")
    response = api_get(session, "/api/v3/settingspolicies", params={"only_defaults": "true"})
    if response.status_code == 200:
        policies = [p for p in unwrap_list_response(response.json()) if isinstance(p, dict)]
        print(f"  Default settings policies: {len(policies)}")
        return policies, "available"

    print_api_warning("default settings policies", response)
    return [], f"HTTP {response.status_code}"


def collect_placement_policies(session: requests.Session) -> tuple[list[dict[str, Any]], str]:
    print("\n-- Placement / Policy Endpoint ----------------------")
    response = api_get(session, "/api/v3/policies")
    if response.status_code == 200:
        policies = [p for p in unwrap_list_response(response.json()) if isinstance(p, dict)]
        print(f"  Total policies: {len(policies)}")
        vmware_count = sum(1 for p in policies if policy_entity_type(p) in VMWARE_ENTITY_TYPES)
        print(f"  VMware-specific by entityType: {vmware_count}")
        print(f"  Other entity types: {len(policies) - vmware_count}")
        return policies, "available"

    print_api_warning("placement/policy endpoint", response)
    return [], f"HTTP {response.status_code}"


# -----------------------------------------------------------------------------
# Classification and analysis
# -----------------------------------------------------------------------------


def classify_settings_policy(
    policy: dict[str, Any],
    default_policy: bool,
    enabled: bool,
    scope_analysis: dict[str, Any],
    default_compare: dict[str, Any],
    action_modes: list[dict[str, Any]],
    audit_available: bool,
    in_audit: bool,
    action_status: dict[str, Any],
    conflict_count: int = 0,
) -> tuple[str, str, list[str]]:
    name = policy_name(policy)
    reasons: list[str] = []

    scope_health = scope_analysis.get("scope_health")
    changed_count = default_compare.get("changed_count")
    age = policy_age_days(policy)
    recent_actions_available = bool(action_status.get("available"))
    has_recent_actions = action_status.get("has_recent_actions")
    high_action_modes = [m for m in action_modes if m.get("risk") == "HIGH"]

    if default_policy:
        reasons.append("default policy")
        return "KEEP", "Default automation/settings policy", reasons

    if not enabled:
        reasons.append("disabled")

    if scope_health in {"EMPTY", "UNRESOLVED", "EMPTY_GROUP", "PARTIAL"}:
        reasons.append(f"scope health: {scope_health}")

    if default_compare.get("status") == "COMPARED":
        if changed_count == 0:
            reasons.append("no detected differences from default policy")
        else:
            reasons.append(f"{changed_count} setting(s) differ from default")
    elif default_compare.get("status") == "NO_DEFAULT_AVAILABLE":
        reasons.append("default comparison unavailable")

    if audit_available:
        if not in_audit:
            reasons.append(f"not found in audit log in last {INACTIVITY_DAYS} days")
    else:
        reasons.append("audit log unavailable; not used as stale evidence")

    if recent_actions_available:
        if not has_recent_actions:
            reasons.append(f"no recent actions found in last {INACTIVITY_DAYS} days")
    else:
        reasons.append("action check unavailable; not used as stale evidence")

    if age is not None and age > 365:
        reasons.append(f"last modified {age} days ago")

    if high_action_modes:
        reasons.append(f"{len(high_action_modes)} AUTOMATIC action mode setting(s)")

    if conflict_count:
        reasons.append(f"{conflict_count} conflicting setting(s) on same scope")

    name_is_test = has_name_pattern(name, TEST_NAME_PATTERNS)
    name_is_orphan = has_name_pattern(name, ORPHAN_PATTERNS)

    # Conservative delete candidate: only when multiple strong indicators agree.
    if (
        not enabled
        and scope_health in {"EMPTY", "UNRESOLVED", "EMPTY_GROUP"}
        and (name_is_test or name_is_orphan or changed_count == 0)
    ):
        return (
            "CANDIDATE_DELETE",
            "Disabled non-default policy with empty/unresolved scope and weak/custom-test signal",
            reasons,
        )

    if not enabled:
        return "REVIEW", "Disabled non-default policy; verify whether it is intentionally retained", reasons

    if scope_health in {"EMPTY", "UNRESOLVED", "EMPTY_GROUP", "PARTIAL"}:
        return "REVIEW", "Scope is empty, unresolved, or partially resolved", reasons

    if changed_count == 0 and default_compare.get("status") == "COMPARED":
        return "REVIEW", "Custom policy appears identical to its default policy", reasons

    if conflict_count:
        return "REVIEW", "Possible overlapping/conflicting policy settings on same scope", reasons

    if high_action_modes:
        return "REVIEW", "Policy contains AUTOMATIC action modes; validate operational intent", reasons

    if age is not None and age > 365 and audit_available and not in_audit:
        return "REVIEW", "Old policy with no audit activity in the checked period", reasons

    return "KEEP", "No strong stale indicators detected", reasons


def extract_group_ref_from_placement(policy: dict[str, Any], keys: tuple[str, ...]) -> Any:
    for key in keys:
        value = policy.get(key)
        if value:
            return value
    return None


def analyze_placement_policy(
    session: requests.Session,
    policy: dict[str, Any],
    group_cache: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    enabled = policy_enabled(policy)
    name = policy_name(policy)
    pid = policy_id(policy)
    etype = policy_entity_type(policy)

    consumer_ref = extract_group_ref_from_placement(policy, ("consumerGroup", "consumerGroupUuid", "consumer"))
    provider_ref = extract_group_ref_from_placement(policy, ("providerGroup", "providerGroupUuid", "provider"))
    merge_groups = policy.get("mergeGroups") or []

    group_refs = []
    for label, ref in (("consumer", consumer_ref), ("provider", provider_ref)):
        if ref:
            group_refs.append((label, ref))
    if isinstance(merge_groups, list):
        for ref in merge_groups:
            group_refs.append(("merge", ref))

    group_details = []
    unresolved = 0
    empty = 0
    for label, ref in group_refs:
        gid = extract_ref_uuid(ref)
        ref_name = extract_ref_name(ref)
        if not gid:
            unresolved += 1
            group_details.append(f"{label}: unresolved ref {ref_name}")
            continue
        group = fetch_group(session, gid, group_cache)
        if not group.get("resolved"):
            unresolved += 1
            group_details.append(f"{label}: unresolved {gid}")
        else:
            member_count = group.get("member_count")
            if member_count == 0:
                empty += 1
            group_details.append(
                f"{label}: {group.get('display_name') or gid} members={member_count if member_count is not None else 'N/A'}"
            )

    reasons = []
    if not enabled:
        reasons.append("disabled")
    if not group_refs:
        reasons.append("no consumer/provider/merge group references detected")
    if unresolved:
        reasons.append(f"{unresolved} unresolved group reference(s)")
    if empty:
        reasons.append(f"{empty} group reference(s) with zero members")

    if not enabled and (has_name_pattern(name, TEST_NAME_PATTERNS) or has_name_pattern(name, ORPHAN_PATTERNS)):
        classification = "CANDIDATE_DELETE"
        justification = "Disabled placement/policy entry with test/orphan-like name"
    elif not enabled or unresolved or empty or not group_refs:
        classification = "REVIEW"
        justification = "Placement policy requires group/reference validation"
    else:
        classification = "INFO"
        justification = "Placement/policy endpoint entry collected for inventory"

    return {
        "classification": classification,
        "justification": justification,
        "source": "placement_policy",
        "id": pid,
        "name": name,
        "entity_type": etype,
        "enabled": enabled,
        "policy_type": normalize_string(policy.get("type") or policy.get("policyType")),
        "consumer_group": extract_ref_name(consumer_ref),
        "provider_group": extract_ref_name(provider_ref),
        "merge_groups": len(merge_groups) if isinstance(merge_groups, list) else 0,
        "unresolved_groups": unresolved,
        "empty_groups": empty,
        "group_details": "; ".join(group_details),
        "reasons": "; ".join(reasons),
        "raw_hash": json_hash(policy),
    }


def build_conflict_index(policy_analyses: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], dict[str, int]]:
    """Detect same entityType + same exact scope set + same setting with different values."""
    buckets: dict[tuple[str, tuple[str, ...], str], list[dict[str, Any]]] = defaultdict(list)

    for analysis in policy_analyses:
        policy = analysis["raw_policy"]
        settings = flatten_policy_settings(policy)
        scopes = tuple(scope_ids(policy))
        etype = policy_entity_type(policy)
        for key, setting in settings.items():
            buckets[(etype, scopes, key)].append(
                {
                    "policy_id": policy_id(policy),
                    "policy_name": policy_name(policy),
                    "entity_type": etype,
                    "scope_ids": scopes,
                    "setting_key": key,
                    "setting_name": setting.get("display_name"),
                    "value": setting.get("value"),
                }
            )

    conflicts = []
    conflict_counts: dict[str, int] = defaultdict(int)
    for (etype, scopes, setting_key), rows in buckets.items():
        if len(rows) < 2:
            continue
        values = {json.dumps(r.get("value"), sort_keys=True, default=str) for r in rows}
        if len(values) <= 1:
            continue
        for row in rows:
            conflict_counts[row["policy_id"]] += 1
            conflicts.append(
                {
                    "entity_type": etype,
                    "scope_ids": ",".join(scopes),
                    "setting_key": setting_key,
                    "setting_name": row.get("setting_name"),
                    "policy_id": row.get("policy_id"),
                    "policy_name": row.get("policy_name"),
                    "value": json.dumps(row.get("value"), ensure_ascii=False, default=str),
                }
            )

    return conflicts, conflict_counts


def analyze_settings_policies(
    session: requests.Session,
    policies: list[dict[str, Any]],
    default_settings_by_entity: dict[str, dict[str, dict[str, Any]]],
    default_ids: set[str],
    audit_available: bool,
    policy_ids_in_audit: set[str],
    cutoff_ms: int,
    group_cache: dict[str, dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    preliminary = []
    action_mode_rows = []
    scope_rows = []
    action_cache: dict[str, dict[str, Any]] = {}

    # First pass collects settings/scopes/action modes. Conflict detection needs
    # all policies before final classification.
    for i, policy in enumerate(policies):
        pid = policy_id(policy)
        default_policy = is_default_policy(policy, default_ids)
        enabled = policy_enabled(policy)
        scopes = analyze_scopes(session, policy, default_policy, group_cache)
        default_compare = compare_to_default(policy, default_settings_by_entity)
        modes = extract_action_modes(policy)
        action_mode_rows.extend(modes)

        for detail in scopes.get("scope_details", []) or []:
            scope_rows.append(
                {
                    "policy_id": pid,
                    "policy_name": policy_name(policy),
                    "entity_type": policy_entity_type(policy),
                    "scope_uuid": detail.get("uuid"),
                    "scope_name": detail.get("name"),
                    "resolved": detail.get("resolved"),
                    "member_count": detail.get("member_count"),
                    "group_type": detail.get("group_type"),
                    "class_name": detail.get("class_name"),
                    "health": detail.get("health"),
                    "error": detail.get("error"),
                }
            )

        preliminary.append(
            {
                "raw_policy": policy,
                "default_policy": default_policy,
                "enabled": enabled,
                "scope_analysis": scopes,
                "default_compare": default_compare,
                "action_modes": modes,
            }
        )
        debug(f"  [{i + 1}/{len(policies)}] {policy_name(policy)}")

    conflicts, conflict_counts = build_conflict_index(preliminary)

    results = []
    for item in preliminary:
        policy = item["raw_policy"]
        pid = policy_id(policy)
        action_status = get_recent_policy_action_status(session, pid, cutoff_ms, action_cache)
        in_audit = pid in policy_ids_in_audit if pid else False
        conflict_count = conflict_counts.get(pid, 0)
        classification, justification, reasons = classify_settings_policy(
            policy=policy,
            default_policy=item["default_policy"],
            enabled=item["enabled"],
            scope_analysis=item["scope_analysis"],
            default_compare=item["default_compare"],
            action_modes=item["action_modes"],
            audit_available=audit_available,
            in_audit=in_audit,
            action_status=action_status,
            conflict_count=conflict_count,
        )

        changed_count = item["default_compare"].get("changed_count")
        action_mode_summary = ", ".join(
            sorted({f"{m['value']}({m['risk']})" for m in item["action_modes"]})
        )

        results.append(
            {
                "classification": classification,
                "justification": justification,
                "source": "settings_policy",
                "id": pid,
                "name": policy_name(policy),
                "entity_type": policy_entity_type(policy),
                "vmware_specific": policy_entity_type(policy) in VMWARE_ENTITY_TYPES,
                "enabled": item["enabled"],
                "is_default": item["default_policy"],
                "scope_count": item["scope_analysis"].get("scope_count", 0),
                "scope_health": item["scope_analysis"].get("scope_health"),
                "resolved_scopes": item["scope_analysis"].get("resolved_count", 0),
                "unresolved_scopes": item["scope_analysis"].get("unresolved_count", 0),
                "empty_scope_groups": item["scope_analysis"].get("empty_group_count", 0),
                "default_compare_status": item["default_compare"].get("status"),
                "changed_settings": changed_count if changed_count is not None else "N/A",
                "total_settings": item["default_compare"].get("total_settings"),
                "action_modes": action_mode_summary,
                "conflict_count": conflict_count,
                "audit_check": "FOUND" if audit_available and in_audit else ("NOT_FOUND" if audit_available else "N/A"),
                "actions_check": (
                    "HAS_RECENT_ACTIONS"
                    if action_status.get("available") and action_status.get("has_recent_actions")
                    else ("NO_RECENT_ACTIONS" if action_status.get("available") else "N/A")
                ),
                "actions_note": action_status.get("note"),
                "last_modified": policy.get("lastModified") or policy.get("modifiedTime") or "N/A",
                "age_days": policy_age_days(policy),
                "reasons": "; ".join(reasons),
                "raw_hash": json_hash(policy),
                "snapshot_scope_ids": scope_ids(policy),
                "snapshot_settings_hash": json_hash(flatten_policy_settings(policy)),
            }
        )

    return results, action_mode_rows, scope_rows, conflicts


# -----------------------------------------------------------------------------
# Snapshot handling
# -----------------------------------------------------------------------------


def snapshot_record(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "id": row.get("id"),
        "source": row.get("source"),
        "name": row.get("name"),
        "entity_type": row.get("entity_type"),
        "enabled": row.get("enabled"),
        "classification": row.get("classification"),
        "scope_ids": row.get("snapshot_scope_ids", []),
        "settings_hash": row.get("snapshot_settings_hash"),
        "raw_hash": row.get("raw_hash"),
    }


def write_and_compare_snapshot(
    rows: list[dict[str, Any]],
    output_path: Path,
) -> tuple[Path | None, list[dict[str, Any]]]:
    if args.no_snapshot:
        return None, []

    snapshot_dir = Path(args.snapshot_dir).expanduser() if args.snapshot_dir else output_path.parent / "snapshots"
    snapshot_dir.mkdir(parents=True, exist_ok=True)
    timestamp = now_utc().strftime("%Y%m%d_%H%M%S")
    snapshot_path = snapshot_dir / f"policy_snapshot_{timestamp}.json"

    current_records = [snapshot_record(row) for row in rows if row.get("id")]
    current_by_key = {f"{r['source']}|{r['id']}": r for r in current_records}

    previous_files = sorted(snapshot_dir.glob("policy_snapshot_*.json"))
    previous_file = previous_files[-1] if previous_files else None
    changes: list[dict[str, Any]] = []

    if previous_file and previous_file != snapshot_path:
        try:
            previous_records = json.loads(previous_file.read_text(encoding="utf-8"))
            previous_by_key = {f"{r.get('source')}|{r.get('id')}": r for r in previous_records}

            for key, current in current_by_key.items():
                previous = previous_by_key.get(key)
                if not previous:
                    changes.append(
                        {
                            "change_type": "ADDED",
                            "source": current.get("source"),
                            "id": current.get("id"),
                            "name": current.get("name"),
                            "field": "",
                            "old": "",
                            "new": "",
                        }
                    )
                    continue
                for field in ("name", "entity_type", "enabled", "classification", "scope_ids", "settings_hash", "raw_hash"):
                    if previous.get(field) != current.get(field):
                        changes.append(
                            {
                                "change_type": "CHANGED",
                                "source": current.get("source"),
                                "id": current.get("id"),
                                "name": current.get("name"),
                                "field": field,
                                "old": json.dumps(previous.get(field), ensure_ascii=False, default=str),
                                "new": json.dumps(current.get(field), ensure_ascii=False, default=str),
                            }
                        )

            for key, previous in previous_by_key.items():
                if key not in current_by_key:
                    changes.append(
                        {
                            "change_type": "REMOVED",
                            "source": previous.get("source"),
                            "id": previous.get("id"),
                            "name": previous.get("name"),
                            "field": "",
                            "old": "",
                            "new": "",
                        }
                    )
        except Exception as exc:  # noqa: BLE001
            changes.append(
                {
                    "change_type": "SNAPSHOT_COMPARE_ERROR",
                    "source": "",
                    "id": "",
                    "name": "",
                    "field": "",
                    "old": str(previous_file),
                    "new": str(exc),
                }
            )

    snapshot_path.write_text(json.dumps(current_records, indent=2, ensure_ascii=False), encoding="utf-8")
    return snapshot_path, changes


# -----------------------------------------------------------------------------
# Excel generation
# -----------------------------------------------------------------------------


def style_header(ws, fill_color: str = "1F4E79") -> None:
    fill = PatternFill("solid", fgColor=fill_color)
    font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def set_widths(ws, widths: list[int]) -> None:
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width


def append_table(ws, headers: list[str], rows: list[dict[str, Any]], widths: list[int] | None = None) -> None:
    ws.append(headers)
    style_header(ws)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
        for cell in ws[ws.max_row]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.font = Font(name="Arial", size=9)
    if widths:
        set_widths(ws, widths)


def make_summary_rows(
    vmware_targets: list[dict[str, Any]],
    disconnected_targets: list[dict[str, Any]],
    entity_counts: dict[str, int | str],
    settings_policies: list[dict[str, Any]],
    default_policies: list[dict[str, Any]],
    placement_policies: list[dict[str, Any]],
    results: list[dict[str, Any]],
    audit_available: bool,
    audit_count: int,
    audit_note: str,
    settings_status: str,
    defaults_status: str,
    placement_status: str,
    snapshot_path: Path | None,
    snapshot_changes: list[dict[str, Any]],
) -> list[list[Any]]:
    numeric_entity_total = sum(v for v in entity_counts.values() if isinstance(v, int))
    rows: list[list[Any]] = [
        ["Turbonomic Stale Policy Audit - VMware On-Premises"],
        [],
        ["Host", TURBO_HOST],
        ["Analysis Date", now_utc().strftime("%Y-%m-%d %H:%M UTC")],
        ["Inactivity Threshold", f"{INACTIVITY_DAYS} days"],
        ["Audit Available", audit_available],
        ["Audit Entries", audit_count],
        ["Audit Note", audit_note],
        [],
        ["API Collection"],
        ["Settings Policies Status", settings_status],
        ["Default Policies Status", defaults_status],
        ["Placement Policies Status", placement_status],
        [],
        ["VMware Environment"],
        ["Total vCenter Targets", len(vmware_targets)],
        ["Targets not in accepted state", len(disconnected_targets)],
        ["Total VMware Entities", numeric_entity_total],
    ]
    for etype, count in entity_counts.items():
        rows.append([f"  {etype}", count])

    rows.extend(
        [
            [],
            ["Policy Inventory"],
            ["Settings Policies", len(settings_policies)],
            ["Default Settings Policies", len(default_policies)],
            ["Placement / Policy Endpoint", len(placement_policies)],
            [],
            ["Classification", "Count"],
        ]
    )
    for classification in CLASSIFICATION_ORDER:
        rows.append([classification, sum(1 for r in results if r.get("classification") == classification)])
    rows.extend(
        [
            ["TOTAL", len(results)],
            [],
            ["VMware-specific", sum(1 for r in results if r.get("vmware_specific"))],
            ["Other entity types", sum(1 for r in results if not r.get("vmware_specific"))],
            [],
            ["Snapshot File", str(snapshot_path) if snapshot_path else "N/A"],
            ["Snapshot Changes", len(snapshot_changes)],
        ]
    )
    return rows


def generate_excel_report(
    output_path: Path,
    results: list[dict[str, Any]],
    vmware_targets: list[dict[str, Any]],
    disconnected_targets: list[dict[str, Any]],
    entity_counts: dict[str, int | str],
    settings_policies: list[dict[str, Any]],
    default_policies: list[dict[str, Any]],
    placement_policies: list[dict[str, Any]],
    placement_rows: list[dict[str, Any]],
    action_mode_rows: list[dict[str, Any]],
    scope_rows: list[dict[str, Any]],
    conflict_rows: list[dict[str, Any]],
    snapshot_changes: list[dict[str, Any]],
    snapshot_path: Path | None,
    audit_available: bool,
    audit_count: int,
    audit_note: str,
    settings_status: str,
    defaults_status: str,
    placement_status: str,
) -> Path:
    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "Summary"
    for row in make_summary_rows(
        vmware_targets,
        disconnected_targets,
        entity_counts,
        settings_policies,
        default_policies,
        placement_policies,
        results,
        audit_available,
        audit_count,
        audit_note,
        settings_status,
        defaults_status,
        placement_status,
        snapshot_path,
        snapshot_changes,
    ):
        ws_summary.append(row)
    ws_summary["A1"].font = Font(bold=True, size=14, name="Arial", color="1F4E79")
    ws_summary.column_dimensions["A"].width = 38
    ws_summary.column_dimensions["B"].width = 80

    policy_headers = [
        "classification",
        "justification",
        "source",
        "name",
        "entity_type",
        "vmware_specific",
        "enabled",
        "is_default",
        "scope_count",
        "scope_health",
        "resolved_scopes",
        "unresolved_scopes",
        "empty_scope_groups",
        "default_compare_status",
        "changed_settings",
        "total_settings",
        "action_modes",
        "conflict_count",
        "audit_check",
        "actions_check",
        "actions_note",
        "last_modified",
        "age_days",
        "reasons",
        "id",
    ]
    policy_widths = [18, 50, 18, 50, 22, 14, 10, 10, 12, 16, 14, 16, 16, 22, 16, 14, 30, 14, 14, 18, 20, 24, 10, 70, 38]

    # Individual classification sheets.
    for classification in CLASSIFICATION_ORDER:
        rows = [r for r in results if r.get("classification") == classification]
        ws = wb.create_sheet(sanitize_sheet_name(classification))
        append_table(ws, policy_headers, rows, policy_widths)
        fill = PatternFill("solid", fgColor=CLASSIFICATION_COLORS.get(classification, "FFFFFF"))
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, len(policy_headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    ws_all = wb.create_sheet("All Policies")
    all_rows = sorted(
        results,
        key=lambda r: (CLASSIFICATION_ORDER.index(r.get("classification", "UNKNOWN")) if r.get("classification") in CLASSIFICATION_ORDER else 99, r.get("name", "")),
    )
    append_table(ws_all, policy_headers, all_rows, policy_widths)

    ws_actions = wb.create_sheet("Action Modes")
    append_table(
        ws_actions,
        ["policy_id", "policy_name", "entity_type", "manager", "setting_uuid", "setting_name", "value", "risk"],
        action_mode_rows,
        [38, 50, 22, 30, 38, 45, 20, 12],
    )

    ws_scopes = wb.create_sheet("Scopes")
    append_table(
        ws_scopes,
        ["policy_id", "policy_name", "entity_type", "scope_uuid", "scope_name", "resolved", "member_count", "group_type", "class_name", "health", "error"],
        scope_rows,
        [38, 50, 22, 38, 45, 10, 14, 22, 22, 16, 60],
    )

    ws_placement = wb.create_sheet("Placement Policies")
    append_table(
        ws_placement,
        [
            "classification",
            "justification",
            "source",
            "name",
            "entity_type",
            "enabled",
            "policy_type",
            "consumer_group",
            "provider_group",
            "merge_groups",
            "unresolved_groups",
            "empty_groups",
            "group_details",
            "reasons",
            "id",
        ],
        placement_rows,
        [18, 50, 18, 50, 22, 10, 22, 35, 35, 14, 16, 14, 70, 60, 38],
    )

    ws_conflicts = wb.create_sheet("Conflicts")
    append_table(
        ws_conflicts,
        ["entity_type", "scope_ids", "setting_key", "setting_name", "policy_id", "policy_name", "value"],
        conflict_rows,
        [22, 60, 45, 45, 38, 50, 40],
    )

    ws_snapshot = wb.create_sheet("Snapshot Changes")
    append_table(
        ws_snapshot,
        ["change_type", "source", "id", "name", "field", "old", "new"],
        snapshot_changes,
        [22, 18, 38, 50, 24, 70, 70],
    )

    ws_vmware = wb.create_sheet("VMware Environment")
    ws_vmware.append(["VMware vCenter Targets"])
    ws_vmware.append(["Target Name", "Category", "Type", "Status", "Accepted Status", "UUID"])
    style_header(ws_vmware)
    for target in vmware_targets:
        status = target.get("status", "UNKNOWN")
        ws_vmware.append(
            [
                target.get("displayName") or target.get("name") or "N/A",
                target.get("category", ""),
                target.get("type", ""),
                status,
                is_target_ok(status),
                target.get("uuid", "N/A"),
            ]
        )
        if not is_target_ok(status):
            for cell in ws_vmware[ws_vmware.max_row]:
                cell.fill = PatternFill("solid", fgColor="FFCCCC")
    set_widths(ws_vmware, [45, 20, 20, 20, 18, 38])

    wb.save(output_path)
    return output_path


# -----------------------------------------------------------------------------
# Main
# -----------------------------------------------------------------------------


def main() -> Path:
    print("=" * 70)
    print("  Turbonomic Stale Policy Audit - VMware On-Premises Focus")
    print("=" * 70)

    output_path = resolve_output_path(OUTPUT_FILE)
    session = login()

    vmware_targets, disconnected_targets = analyze_vmware_targets(session)
    entity_counts = get_vmware_entity_counts(session)

    cutoff_date = now_utc() - timedelta(days=INACTIVITY_DAYS)
    cutoff_ms = int(cutoff_date.timestamp() * 1000)

    audit_available, policy_ids_in_audit, audit_count, audit_note = collect_audit_log(session, cutoff_ms)

    settings_policies, settings_status = collect_settings_policies(session)
    default_policies, defaults_status = collect_default_settings_policies(session)
    placement_policies, placement_status = collect_placement_policies(session)

    _, default_settings_by_entity, default_ids = build_default_policy_maps(default_policies)

    group_cache: dict[str, dict[str, Any]] = {}

    settings_results, action_mode_rows, scope_rows, conflict_rows = analyze_settings_policies(
        session=session,
        policies=settings_policies,
        default_settings_by_entity=default_settings_by_entity,
        default_ids=default_ids,
        audit_available=audit_available,
        policy_ids_in_audit=policy_ids_in_audit,
        cutoff_ms=cutoff_ms,
        group_cache=group_cache,
    )

    placement_rows = [analyze_placement_policy(session, p, group_cache) for p in placement_policies]

    # Include placement policy rows in the global view, but keep source-specific
    # details in the Placement Policies sheet.
    placement_as_global = []
    for row in placement_rows:
        placement_as_global.append(
            {
                "classification": row.get("classification"),
                "justification": row.get("justification"),
                "source": row.get("source"),
                "id": row.get("id"),
                "name": row.get("name"),
                "entity_type": row.get("entity_type"),
                "vmware_specific": row.get("entity_type") in VMWARE_ENTITY_TYPES,
                "enabled": row.get("enabled"),
                "is_default": False,
                "scope_count": "N/A",
                "scope_health": "N/A",
                "resolved_scopes": "N/A",
                "unresolved_scopes": row.get("unresolved_groups"),
                "empty_scope_groups": row.get("empty_groups"),
                "default_compare_status": "N/A",
                "changed_settings": "N/A",
                "total_settings": "N/A",
                "action_modes": "N/A",
                "conflict_count": "N/A",
                "audit_check": "N/A",
                "actions_check": "N/A",
                "actions_note": "N/A",
                "last_modified": "N/A",
                "age_days": "N/A",
                "reasons": row.get("reasons"),
                "raw_hash": row.get("raw_hash"),
                "snapshot_scope_ids": [],
                "snapshot_settings_hash": row.get("raw_hash"),
            }
        )

    results = settings_results + placement_as_global

    snapshot_path, snapshot_changes = write_and_compare_snapshot(results, output_path)

    to_candidate_delete = [r for r in results if r.get("classification") == "CANDIDATE_DELETE"]
    to_review = [r for r in results if r.get("classification") == "REVIEW"]
    to_keep = [r for r in results if r.get("classification") == "KEEP"]
    info_rows = [r for r in results if r.get("classification") == "INFO"]
    unknown_rows = [r for r in results if r.get("classification") == "UNKNOWN"]
    vmware_flagged = [r for r in results if r.get("vmware_specific")]
    non_vmware_flagged = [r for r in results if not r.get("vmware_specific")]

    print(f"\n{'=' * 70}")
    print("  AUDIT SUMMARY")
    print(f"{'=' * 70}")
    print(f"  Host: {TURBO_HOST}")
    print(f"  Inactivity threshold: {INACTIVITY_DAYS} days")
    print(f"  Analysis date: {now_utc().strftime('%Y-%m-%d %H:%M UTC')}")
    print(f"  Audit available: {audit_available} ({audit_note})")
    print(f"{'=' * 70}")
    print(f"  VMware targets    : {len(vmware_targets)} ({len(disconnected_targets)} not accepted)")
    print(f"  VMware entities   : {sum(v for v in entity_counts.values() if isinstance(v, int))}")
    print(f"{'=' * 70}")
    print(f"  TOTAL rows        : {len(results)}")
    print(f"    VMware-specific : {len(vmware_flagged)}")
    print(f"    Other types     : {len(non_vmware_flagged)}")
    print(f"{'=' * 70}")
    print(f"  CANDIDATE_DELETE  : {len(to_candidate_delete)}")
    print(f"  REVIEW            : {len(to_review)}")
    print(f"  KEEP              : {len(to_keep)}")
    print(f"  INFO              : {len(info_rows)}")
    print(f"  UNKNOWN           : {len(unknown_rows)}")
    print(f"{'=' * 70}\n")

    generate_excel_report(
        output_path=output_path,
        results=results,
        vmware_targets=vmware_targets,
        disconnected_targets=disconnected_targets,
        entity_counts=entity_counts,
        settings_policies=settings_policies,
        default_policies=default_policies,
        placement_policies=placement_policies,
        placement_rows=placement_rows,
        action_mode_rows=action_mode_rows,
        scope_rows=scope_rows,
        conflict_rows=conflict_rows,
        snapshot_changes=snapshot_changes,
        snapshot_path=snapshot_path,
        audit_available=audit_available,
        audit_count=audit_count,
        audit_note=audit_note,
        settings_status=settings_status,
        defaults_status=defaults_status,
        placement_status=placement_status,
    )

    print(f"Report exported: {output_path.resolve()}")
    if snapshot_path:
        print(f"Snapshot saved:  {snapshot_path.resolve()}")
    print(
        "Sheets: 'Summary' | 'CANDIDATE_DELETE' | 'REVIEW' | 'KEEP' | 'INFO' | "
        "'UNKNOWN' | 'All Policies' | 'Action Modes' | 'Scopes' | "
        "'Placement Policies' | 'Conflicts' | 'Snapshot Changes' | 'VMware Environment'"
    )
    return output_path


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("Interrupted by user")
        sys.exit(130)
